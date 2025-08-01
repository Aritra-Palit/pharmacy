from datetime import datetime
from flask import Flask, redirect, render_template, request, jsonify, send_from_directory
from openpyxl import load_workbook
import os
from werkzeug.exceptions import HTTPException

app = Flask(__name__)

UPLOAD_FOLDER = os.path.join('static', 'profile_pics')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/view-bills')
def view_bills_page():
    return render_template('viewBills.html')

@app.route('/my-profile')
def profile():
    return render_template('profile.html')

@app.errorhandler(HTTPException)
def handle_http_errors(e):
    return redirect('/')

@app.errorhandler(Exception)
def handle_all_errors(e):
    return redirect('/')

@app.route('/add-medicine')
def add_medicine_page():
    return render_template('addMedicine.html')

@app.route('/view-medicine')
def view_medicine_page():
    return render_template('viewMedicine.html')

@app.route('/upload-profile', methods=['POST'])
def upload_profile():
    file = request.files.get('image')
    user_id = request.form.get('user_id')  # Assume user is identified
    if not file or not user_id:
        return 'Missing data', 400

    filename = f"{user_id}.jpg"
    file.save(os.path.join(UPLOAD_FOLDER, filename))
    return 'Uploaded', 200

@app.route('/profile-picture/<user_id>')
def view_profile(user_id):
    filename = f"{user_id}.jpg"
    path = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(path):
        return send_from_directory(UPLOAD_FOLDER, filename)
    return 'Not Found', 404

@app.route('/delete-profile/<user_id>', methods=['DELETE'])
def delete_profile(user_id):
    path = os.path.join(UPLOAD_FOLDER, f"{user_id}.jpg")
    if os.path.exists(path):
        os.remove(path)
        return 'Deleted', 200
    return 'Not Found', 404

@app.route('/login', methods=['POST'])
def login():
    from openpyxl import load_workbook

    username = request.form.get('username', '').strip().lower()
    password = request.form.get('password', '').strip()

    excel_path = os.path.join(os.path.dirname(__file__), 'master.xlsx')
    wb = load_workbook(excel_path)
    if 'Usernames' not in wb.sheetnames:
        return "Login sheet not found", 400

    ws = wb['Usernames']
    for row in ws.iter_rows(min_row=2, values_only=True):
        user, pwd, name, emp_id, emp_Type = row[:5]
        if str(user).strip().lower() == username and str(pwd).strip() == password:
            return {
                "name": name,
                "employeeId": emp_id,
                "EmpType" : emp_Type
            }, 200

    return "Invalid credentials", 401

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/add-customer')
def add_customer_page():
    return render_template('addCustomer.html')

@app.route('/add-customer', methods=['POST'])
def add_customer():
    phone = request.form.get('phone')
    name = request.form.get('name')
    address = request.form.get('address')

    

    excel_path = os.path.join(os.path.dirname(__file__), 'master.xlsx')

    try:
        wb = load_workbook(excel_path)
        sheet_name = 'Customer'
        if sheet_name not in wb.sheetnames:
            return f"Sheet '{sheet_name}' not found", 400

        ws = wb[sheet_name]
        ws.append([phone, name, address])
        wb.save(excel_path)
        return ('', 204)

    except Exception as e:
        print("❗ Error:", str(e))
        return "Error: " + str(e), 500

@app.route('/customers')
def get_customers():
    excel_path = os.path.join(os.path.dirname(__file__), 'master.xlsx')
    wb = load_workbook(excel_path)
    sheet_name = 'Customer'
    if sheet_name not in wb.sheetnames:
        return jsonify([])

    ws = wb[sheet_name]
    customers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        customer = {
            "Phone": str(row[0]).strip(),
            "Name": str(row[1]).strip(),
            "Address": str(row[2]).strip()
        }
        customers.append(customer)

    return jsonify(customers)


@app.route('/medicines')
def get_medicines():
    excel_path = os.path.join(os.path.dirname(__file__), 'master.xlsx')
    wb = load_workbook(excel_path)
    sheet_name = 'Medicines'
    if sheet_name not in wb.sheetnames:
        return jsonify([])

    ws = wb[sheet_name]
    medicines = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        medicine = {
            "MedicineID": row[0],
            "name": row[1],
            "Price": row[3],
            "Quantity": row[2],
            "dateAdded": str(row[4]),
            "lastUpdated": str(row[5])
        }
        print(str(row[5]))
        medicines.append(medicine)

    return jsonify(medicines)

@app.route("/add-medicine", methods=["POST"])
def add_medicine():
    try:
        data = request.get_json()
        excel_path = os.path.join(os.path.dirname(__file__), "master.xlsx")
        wb = load_workbook(excel_path)
        ws = wb["Medicines"]

        # Format dates
        from datetime import datetime
        date_added = datetime.strptime(data["dateAdded"], "%Y-%m-%d").strftime("%d-%m-%Y")
        last_updated = datetime.strptime(data["lastUpdated"], "%Y-%m-%d").strftime("%d-%m-%Y")

        found = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            if str(row[0].value).strip().lower() == data["medicineID"].strip().lower():
                # ✅ Update quantity and lastUpdated
                existing_qty = int(row[2].value)
                added_qty = int(data["quantity"])
                row[2].value = existing_qty + added_qty
                row[3].value = float(data["price"])  # update price if needed
                row[5].value = last_updated
                found = True
                break

        if not found:
            # New entry
            ws.append([
                data["medicineID"],
                data["name"],
                int(data["quantity"]),
                float(data["price"]),
                date_added,
                last_updated
            ])

        wb.save(excel_path)
        return '', 204

    except Exception as e:
        print("❌ Error:", e)
        return f"Internal Server Error: {e}", 500


@app.route('/process-bill', methods=['POST'])
def process_bill():
    try:
        data = request.get_json()

        # 🔹 Extract billing data
        user_id = data.get('userId')
        user_name = data.get('userName')
        phone = data.get('customerPhone')
        name = data.get('customerName')
        address = data.get('customerAddress')
        total_amount = data.get('total_amount', 0)
        products = data.get('products', [])       # [{product, price}]
        items = data.get('items', [])             # [{id, quantity}]

        excel_path = os.path.join(os.path.dirname(__file__), 'master.xlsx')
        wb = load_workbook(excel_path)

        # 🔸 Deduct stock
        if 'Medicines' not in wb.sheetnames:
            return jsonify({"error": "Medicines sheet missing"}), 500

        ws_meds = wb['Medicines']
        med_headers = {cell.value: idx for idx, cell in enumerate(next(ws_meds.iter_rows(min_row=1, max_row=1)))}

        for item in items:
            med_id = item.get('id')
            qty_to_deduct = int(item.get('quantity', 0))

            if not med_id or qty_to_deduct <= 0:
                continue

            for row in ws_meds.iter_rows(min_row=2):
                if str(row[med_headers['MedicineID']].value) == med_id:
                    qty_cell = row[med_headers['Quantity']]
                    current_qty = int(qty_cell.value or 0)
                    qty_cell.value = max(0, current_qty - qty_to_deduct)
                    break

        # 🔸 Save bill
        if 'Bills' not in wb.sheetnames:
            ws_bills = wb.create_sheet('Bills')
            ws_bills.append([
                "Timestamp", "User ID", "User Name", "Customer Phone", "Customer Name",
                "Customer Address", "Total Amount", "Product", "Price"
            ])
        else:
            ws_bills = wb['Bills']

        timestamp = datetime.now().strftime('%d-%m-%Y %H:%M:%S')

        if products:
            first = products[0]
            ws_bills.append([
                timestamp, user_id, user_name, phone, name,
                address, total_amount, first['product'], first['price']
            ])
            for item in products[1:]:
                ws_bills.append(["", "", "", "", "", "", "", item['product'], item['price']])
        else:
            ws_bills.append([
                timestamp, user_id, user_name, phone, name,
                address, total_amount, "", ""
            ])

        # ✅ Save everything
        wb.save(excel_path)
        return '', 204

    except Exception as e:
        print("❌ Error in process_bill:", str(e))
        return jsonify({"error": str(e)}), 500
    
@app.route('/api/bills')
def get_bills():
    excel_path = os.path.join(os.path.dirname(__file__), 'master.xlsx')
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Bills']

    bills = []
    current = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        timestamp, user_id, user_name, phone, name, address, total, product, price = row

        if timestamp:  # New bill
            current = {
                "timestamp": timestamp,
                "user": user_name,
                "customer": name,
                "phone": phone,
                "address": address,
                "total": total,
                "products": []
            }
            if product:
                current["products"].append({"product": product, "price": price})
            bills.append(current)

        else:  # Part of same bill
            if product and current:
                current["products"].append({"product": product, "price": price})

    return jsonify(bills)

@app.route('/update-profile', methods=['POST'])
def update_profile():
    try:
        data = request.get_json()
        emp_id = data.get('empId')
        name = data.get('name')
        username = data.get('username')
        password = data.get('password')

        excel_path = os.path.join(os.path.dirname(__file__), 'master.xlsx')
        wb = load_workbook(excel_path)
        ws = wb['Usernames']

        # Find row with emp_id
        headers = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}
        for row in ws.iter_rows(min_row=2):
            if str(row[headers['Emp_ID']].value) == emp_id:
                row[headers['Name']].value = name
                row[headers['Username']].value = username
                if password:
                    row[headers['Password']].value = password
                break

        wb.save(excel_path)
        return '', 204

    except Exception as e:
        print("❌ Error updating profile:", str(e))
        return jsonify({"error": str(e)}), 500



if __name__ == '__main__':
    app.run(debug=True)
